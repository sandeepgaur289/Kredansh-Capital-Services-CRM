import os,re,csv,io,json,sqlite3,secrets,shutil,datetime
from pathlib import Path
from functools import wraps
from urllib.parse import quote
from flask import Flask,request,redirect,url_for,session,render_template,flash,jsonify,send_file,abort
from werkzeug.security import generate_password_hash,check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import load_workbook

BASE=Path(__file__).resolve().parent
DATABASE_URL=os.environ.get('DATABASE_URL','').strip()
IS_POSTGRES=DATABASE_URL.startswith('postgres://') or DATABASE_URL.startswith('postgresql://')
DB_PATH=BASE/'kredansh_online.db'; SEED_DB=BASE/'seed/kredansh_erp.db'
UPLOAD_DIR=Path(os.environ.get('UPLOAD_DIR',str(BASE/'uploads'))); UPLOAD_DIR.mkdir(parents=True,exist_ok=True)
app=Flask(__name__); app.secret_key=os.environ.get('KREDANSH_SECRET') or secrets.token_hex(32)
app.config['MAX_CONTENT_LENGTH']=int(os.environ.get('MAX_UPLOAD_MB','30'))*1024*1024
if os.environ.get('RAILWAY_ENVIRONMENT') or os.environ.get('CLOUD_MODE','').lower() in ('1','true','yes'):
    app.config.update(SESSION_COOKIE_HTTPONLY=True,SESSION_COOKIE_SAMESITE='Lax',SESSION_COOKIE_SECURE=True)
COLLECTION_STATUSES=['Pending','Contacted','PTP','Broken PTP','Part Paid','Paid','Legal','Skip','Not Traceable','Closed']
ENFORCEMENT_STATUSES=['Pending Status','DM Order Pending','DM Order Available','Chaspa Affixed','Possession Scheduled','Rescheduled','Second Attempt','Possession Done','Settled','Regularized','Hold by Bank','Cancelled']
ROLES=['Admin','Manager','Recovery Officer','Data Entry','Viewer']

def now(): return datetime.datetime.now().isoformat(timespec='seconds')
def today(): return datetime.date.today().isoformat()
def money(v):
    try:return '₹'+format(float(v or 0),',.0f')
    except:return '₹0'

class Cur:
    def __init__(self,c,lastrowid=None): self.c=c; self.lastrowid=lastrowid
    def fetchone(self): return self.c.fetchone()
    def fetchall(self): return self.c.fetchall()
class PG:
    def __init__(self,url):
        import psycopg
        from psycopg.rows import dict_row
        self.con=psycopg.connect(url,row_factory=dict_row,connect_timeout=15)
    def execute(self,sql,p=()):
        sql=sql.replace('?','%s'); cur=self.con.cursor(); cur.execute(sql,p or ()); return Cur(cur)
    def commit(self): self.con.commit()
    def rollback(self): self.con.rollback()
    def close(self): self.con.close()
def db():
    if IS_POSTGRES:return PG(DATABASE_URL)
    c=sqlite3.connect(DB_PATH,timeout=30); c.row_factory=sqlite3.Row
    try:c.execute('pragma journal_mode=WAL');c.execute('pragma busy_timeout=30000')
    except:pass
    return c

def init_db():
    if not IS_POSTGRES and not DB_PATH.exists() and SEED_DB.exists(): shutil.copy2(SEED_DB,DB_PATH)
    c=db()
    if IS_POSTGRES:
        stmts=[
        "CREATE TABLE IF NOT EXISTS users(id BIGSERIAL PRIMARY KEY,username TEXT UNIQUE NOT NULL,password TEXT NOT NULL,full_name TEXT,role TEXT,status TEXT DEFAULT 'Active',last_login TEXT)",
        "CREATE TABLE IF NOT EXISTS banks(id BIGSERIAL PRIMARY KEY,bank_name TEXT UNIQUE NOT NULL,contact_person TEXT,mobile TEXT,email TEXT,address TEXT,empanelment_no TEXT,agreement_date TEXT,status TEXT DEFAULT 'Active',remarks TEXT)",
        "CREATE TABLE IF NOT EXISTS employees(id BIGSERIAL PRIMARY KEY,employee_code TEXT UNIQUE NOT NULL,employee_name TEXT NOT NULL,designation TEXT,mobile TEXT,email TEXT,branch TEXT,area TEXT,district TEXT,department TEXT,status TEXT DEFAULT 'Active',alternate_mobile TEXT,date_of_birth TEXT,date_of_joining TEXT,blood_group TEXT,emergency_contact TEXT,address TEXT,aadhaar_number TEXT,pan_number TEXT,driving_licence_number TEXT,bank_name TEXT,bank_account TEXT,ifsc TEXT,reporting_manager TEXT,photo_path TEXT,remarks TEXT)",
        "CREATE TABLE IF NOT EXISTS collection_cases(id BIGSERIAL PRIMARY KEY,bank_id BIGINT,loan_number TEXT UNIQUE NOT NULL,customer_name TEXT NOT NULL,area TEXT,district TEXT,mobile TEXT,alternate_mobile TEXT,loan_amount DOUBLE PRECISION DEFAULT 0,emi DOUBLE PRECISION DEFAULT 0,overdue_amount DOUBLE PRECISION DEFAULT 0,total_outstanding DOUBLE PRECISION DEFAULT 0,present_address TEXT,permanent_address TEXT,employee_id BIGINT,allocation_date TEXT,ptp_date TEXT,ptp_amount DOUBLE PRECISION DEFAULT 0,amount_collected DOUBLE PRECISION DEFAULT 0,collection_date TEXT,receipt_number TEXT,payment_mode TEXT,status TEXT DEFAULT 'Pending',next_follow_up TEXT,remarks TEXT,created_at TEXT,updated_at TEXT,visit_status TEXT,visit_date TEXT)",
        "CREATE TABLE IF NOT EXISTS enforcement_cases(id BIGSERIAL PRIMARY KEY,case_number TEXT UNIQUE NOT NULL,bank_id BIGINT,loan_number TEXT,borrower_name TEXT NOT NULL,village_city TEXT,tehsil TEXT,district TEXT,mobile TEXT,property_address TEXT,bank_authorised_officer TEXT,dm_order_status TEXT,dm_order_date TEXT,duty_magistrate TEXT,field_officer TEXT,employee_id BIGINT,allocation_date TEXT,confirmation_date TEXT,affiliation_date TEXT,cost_approval_status TEXT,administrative_expense DOUBLE PRECISION DEFAULT 0,recovery_amount DOUBLE PRECISION DEFAULT 0,possession_date TEXT,second_attempt_date TEXT,next_follow_up TEXT,status TEXT DEFAULT 'Pending Status',remarks TEXT,created_at TEXT,updated_at TEXT)",
        "CREATE TABLE IF NOT EXISTS tehsil_police_contacts(id BIGSERIAL PRIMARY KEY,state TEXT,district TEXT,tehsil TEXT,sub_tehsil TEXT,tehsildar_name TEXT,tehsildar_contact TEXT,reader_name TEXT,reader_contact TEXT,police_station_name TEXT,sho_name TEXT,sho_contact TEXT,police_station_contact TEXT,office_address TEXT,email TEXT,status TEXT DEFAULT 'Active',remarks TEXT,created_at TEXT,updated_at TEXT)",
        "CREATE TABLE IF NOT EXISTS activity_log(id BIGSERIAL PRIMARY KEY,module TEXT,record_id BIGINT,action_date TEXT,action TEXT,details TEXT,employee_id BIGINT,created_by TEXT,created_at TEXT)",
        "CREATE TABLE IF NOT EXISTS collection_allocations(id BIGSERIAL PRIMARY KEY,collection_case_id BIGINT,allocation_month TEXT,allocation_batch TEXT,bank_id BIGINT,employee_id BIGINT,allocation_date TEXT,allocated_outstanding DOUBLE PRECISION DEFAULT 0,allocated_overdue DOUBLE PRECISION DEFAULT 0,collected_amount DOUBLE PRECISION DEFAULT 0,allocation_status TEXT,fresh_repeat TEXT,opening_case_status TEXT,closing_case_status TEXT,closed_date TEXT,close_reason TEXT,created_at TEXT,updated_at TEXT)",
        "CREATE TABLE IF NOT EXISTS collection_payments(id BIGSERIAL PRIMARY KEY,collection_case_id BIGINT,allocation_id BIGINT,payment_date TEXT,amount DOUBLE PRECISION DEFAULT 0,payment_mode TEXT,receipt_number TEXT,remarks TEXT,employee_id BIGINT,created_at TEXT,created_by TEXT)",
        "CREATE TABLE IF NOT EXISTS collection_visits(id BIGSERIAL PRIMARY KEY,collection_case_id BIGINT,allocation_id BIGINT,visit_date TEXT,visit_result TEXT,person_met TEXT,remarks TEXT,employee_id BIGINT,created_at TEXT,created_by TEXT)",
        "CREATE TABLE IF NOT EXISTS collection_ptp_ledger(id BIGSERIAL PRIMARY KEY,collection_case_id BIGINT,allocation_id BIGINT,ptp_date TEXT,ptp_amount DOUBLE PRECISION DEFAULT 0,outcome TEXT,remarks TEXT,employee_id BIGINT,created_at TEXT,created_by TEXT)"
        ]
        for s in stmts:c.execute(s)
    row=c.execute('select count(*) c from users').fetchone(); count=row['c']
    if not count:c.execute('insert into users(username,password,full_name,role,status) values(?,?,?,?,?)',('admin',generate_password_hash('admin123'),'Administrator','Admin','Active'))
    c.commit();c.close()
    if IS_POSTGRES:migrate_seed()
def migrate_seed():
    if not SEED_DB.exists():return
    c=db(); row=c.execute('select count(*) c from banks').fetchone()
    if row['c']>0:c.close();return
    s=sqlite3.connect(SEED_DB);s.row_factory=sqlite3.Row
    tables=['banks','employees','collection_cases','enforcement_cases','tehsil_police_contacts','collection_allocations','collection_payments','collection_visits','collection_ptp_ledger','activity_log']
    for table in tables:
        try:rows=s.execute(f'select * from {table}').fetchall()
        except:continue
        for r in rows:
            cols=[k for k in r.keys() if k!='id']; vals=[r[k] for k in cols]
            try:c.execute(f"insert into {table}({','.join(cols)}) values({','.join('?' for _ in cols)})",vals);c.commit()
            except:c.rollback()
    try:
        for r in s.execute('select * from users').fetchall():
            pw=r['password'] or ''; pw=pw if pw.startswith(('pbkdf2:','scrypt:')) else generate_password_hash(pw)
            try:c.execute('insert into users(username,password,full_name,role,status,last_login) values(?,?,?,?,?,?)',(r['username'],pw,r['full_name'],r['role'],r['status'],r['last_login']));c.commit()
            except:c.rollback()
    except:pass
    s.close();c.close()

def auth(fn):
    @wraps(fn)
    def w(*a,**k):
        if not session.get('uid'):return redirect(url_for('login',next=request.path))
        return fn(*a,**k)
    return w
def mgr(fn):
    @wraps(fn)
    def w(*a,**k):
        if session.get('role') not in ('Admin','Manager'):abort(403)
        return fn(*a,**k)
    return w
def log(module,rid,action,details=''):
    c=db();c.execute('insert into activity_log(module,record_id,action_date,action,details,created_by,created_at) values(?,?,?,?,?,?,?)',(module,rid,today(),action,details,session.get('username','system'),now()));c.commit();c.close()
@app.context_processor
def context():return dict(money=money)

@app.route('/healthz')
def healthz():return {'status':'ok','app':'Kredansh Capital Online ERP'}
@app.route('/system/health')
def syshealth():return {'ok':True,'postgres':IS_POSTGRES}
@app.route('/manifest.webmanifest')
def manifest():return app.response_class(json.dumps({'name':'KREDANSH CAPITAL ERP','short_name':'KREDANSH','start_url':'/','scope':'/','display':'standalone','background_color':'#07111f','theme_color':'#07111f','icons':[{'src':'/static/pwa/icon-192.png','sizes':'192x192','type':'image/png'},{'src':'/static/pwa/icon-512.png','sizes':'512x512','type':'image/png'}]}),mimetype='application/manifest+json')
@app.route('/service-worker.js')
def worker():
    return app.response_class("const C='kredansh-v1';self.addEventListener('install',e=>self.skipWaiting());self.addEventListener('activate',e=>e.waitUntil(self.clients.claim()));",mimetype='application/javascript')

@app.route('/login',methods=['GET','POST'])
def login():
    if request.method=='POST':
        u=request.form.get('username','').strip();p=request.form.get('password','');c=db();r=c.execute('select * from users where username=?',(u,)).fetchone()
        if r and r['status']!='Inactive':
            stored=r['password'] or ''; ok=check_password_hash(stored,p) if stored.startswith(('pbkdf2:','scrypt:')) else secrets.compare_digest(stored,p)
            if ok:
                session.update(uid=r['id'],username=r['username'],full_name=r['full_name'],role=r['role'])
                if not stored.startswith(('pbkdf2:','scrypt:')):c.execute('update users set password=? where id=?',(generate_password_hash(p),r['id']))
                c.execute('update users set last_login=? where id=?',(now(),r['id']));c.commit();c.close();return redirect(request.args.get('next') or '/')
        c.close();flash('Invalid Login ID or Password','danger')
    return render_template('login.html')
@app.route('/logout')
def logout():session.clear();return redirect('/login')

@app.route('/')
@auth
def dashboard():
    c=db()
    def sc(q,p=()):return c.execute(q,p).fetchone()['c']
    k={'Banks':sc("select count(*) c from banks where coalesce(status,'Active')='Active'"),'Employees':sc("select count(*) c from employees where coalesce(status,'Active')='Active'"),'Collection Cases':sc('select count(*) c from collection_cases'),'Enforcement Cases':sc('select count(*) c from enforcement_cases'),'PTP Today':sc('select count(*) c from collection_cases where ptp_date=?',(today(),)),'Possession Today':sc('select count(*) c from enforcement_cases where possession_date=?',(today(),)),'Follow-ups Today':sc('select (select count(*) from collection_cases where next_follow_up=?)+(select count(*) from enforcement_cases where next_follow_up=?) c',(today(),today()))}
    out=c.execute('select coalesce(sum(total_outstanding),0) c from collection_cases').fetchone()['c'];collected=c.execute('select coalesce(sum(amount_collected),0) c from collection_cases').fetchone()['c'];k['Outstanding']=out;k['Collected']=collected
    enf=c.execute("select b.bank_name,count(e.id) total,sum(case when e.possession_date is not null and e.possession_date<>'' then 1 else 0 end) possession_date,sum(case when e.status='Chaspa Affixed' then 1 else 0 end) chaspa,sum(case when e.status='Settled' then 1 else 0 end) settled,sum(case when e.status='Possession Done' then 1 else 0 end) possession_done,sum(case when e.status='Rescheduled' then 1 else 0 end) rescheduled,sum(case when e.status='Hold by Bank' then 1 else 0 end) hold_bank,sum(case when e.status='DM Order Pending' then 1 else 0 end) dm_pending,sum(case when e.status='Regularized' then 1 else 0 end) regularized,sum(case when e.status='Cancelled' then 1 else 0 end) cancelled from banks b left join enforcement_cases e on e.bank_id=b.id group by b.id,b.bank_name order by b.bank_name").fetchall()
    col=c.execute("select b.bank_name,count(x.id) total,coalesce(sum(x.total_outstanding),0) outstanding,coalesce(sum(x.amount_collected),0) collected,sum(case when x.status='Pending' then 1 else 0 end) pending,sum(case when x.status='PTP' then 1 else 0 end) ptp,sum(case when x.status='Broken PTP' then 1 else 0 end) broken_ptp,sum(case when x.status='Part Paid' then 1 else 0 end) part_paid,sum(case when x.status='Paid' then 1 else 0 end) paid,sum(case when x.status='Closed' then 1 else 0 end) closed from banks b left join collection_cases x on x.bank_id=b.id group by b.id,b.bank_name order by b.bank_name").fetchall();c.close();return render_template('dashboard.html',k=k,enf=enf,col=col)

@app.route('/cases/<module>')
@auth
def cases(module):
    if module not in ('collection','enforcement'):abort(404)
    t='collection_cases' if module=='collection' else 'enforcement_cases'; name='customer_name' if module=='collection' else 'borrower_name'; ref='loan_number' if module=='collection' else 'case_number';q=request.args.get('q','');status=request.args.get('status','');bank=request.args.get('bank','');employee=request.args.get('employee','');district=request.args.get('district','');where=['1=1'];p=[]
    if q:where.append(f'(x.{name} like ? or x.{ref} like ? or coalesce(x.mobile,\'\') like ? or coalesce(x.district,\'\') like ?)');p += [f'%{q}%']*4
    if status:where.append('x.status=?');p.append(status)
    if bank:where.append('b.id=?');p.append(bank)
    if employee=='unallocated':where.append('x.employee_id is null')
    elif employee:where.append('e.id=?');p.append(employee)
    if district:where.append('x.district=?');p.append(district)
    c=db();rows=c.execute(f"select x.*,b.bank_name,e.employee_name from {t} x left join banks b on b.id=x.bank_id left join employees e on e.id=x.employee_id where {' and '.join(where)} order by x.updated_at desc,x.id desc",p).fetchall();banks=c.execute("select id,bank_name from banks where coalesce(status,'Active')='Active' order by bank_name").fetchall();emps=c.execute("select id,employee_name from employees where coalesce(status,'Active')='Active' order by employee_name").fetchall();dists=c.execute(f"select distinct district from {t} where coalesce(district,'')<>'' order by district").fetchall();c.close();return render_template('cases.html',module=module,rows=rows,banks=banks,emps=emps,dists=dists,statuses=COLLECTION_STATUSES if module=='collection' else ENFORCEMENT_STATUSES)

@app.route('/case/<module>/<int:rid>')
@auth
def detail(module,rid):
    t='collection_cases' if module=='collection' else 'enforcement_cases';c=db();r=c.execute(f'select x.*,b.bank_name,e.employee_name from {t} x left join banks b on b.id=x.bank_id left join employees e on e.id=x.employee_id where x.id=?',(rid,)).fetchone();hist=c.execute('select * from activity_log where module=? and record_id=? order by id desc',(module,rid)).fetchall();pays=visits=ptps=[]
    if module=='collection':pays=c.execute('select * from collection_payments where collection_case_id=? order by id desc',(rid,)).fetchall();visits=c.execute('select * from collection_visits where collection_case_id=? order by id desc',(rid,)).fetchall();ptps=c.execute('select * from collection_ptp_ledger where collection_case_id=? order by id desc',(rid,)).fetchall()
    c.close();return render_template('detail.html',module=module,r=r,hist=hist,pays=pays,visits=visits,ptps=ptps,statuses=COLLECTION_STATUSES if module=='collection' else ENFORCEMENT_STATUSES)
@app.route('/case/<module>/<int:rid>/save',methods=['POST'])
@auth
def save_case(module,rid):
    t='collection_cases' if module=='collection' else 'enforcement_cases';fields=['status','next_follow_up','remarks','mobile','district','employee_id']+(['ptp_date','ptp_amount','amount_collected','collection_date','visit_status','visit_date'] if module=='collection' else ['dm_order_status','dm_order_date','duty_magistrate','field_officer','confirmation_date','affiliation_date','possession_date','second_attempt_date','cost_approval_status','administrative_expense','recovery_amount']);d={f:request.form.get(f,'').strip() for f in fields};d['employee_id']=d['employee_id'] or None;c=db();c.execute(f"update {t} set "+','.join(f'{f}=?' for f in fields)+',updated_at=? where id=?',list(d.values())+[now(),rid]);c.commit();c.close();log(module,rid,'Case Updated',f"Status {d['status']}");flash('Case updated successfully','success');return redirect(url_for('detail',module=module,rid=rid))
@app.route('/cases/<module>/bulk',methods=['POST'])
@auth
def bulk(module):
    ids=[int(x) for x in request.form.getlist('ids') if x.isdigit()];action=request.form.get('action');t='collection_cases' if module=='collection' else 'enforcement_cases';c=db()
    for rid in ids:
        if action=='status':c.execute(f'update {t} set status=?,updated_at=? where id=?',(request.form.get('status'),now(),rid))
        elif action=='allocate':c.execute(f'update {t} set employee_id=?,allocation_date=?,updated_at=? where id=?',(request.form.get('employee_id') or None,today(),now(),rid))
        elif action=='delete' and session.get('role') in ('Admin','Manager'):c.execute(f'delete from {t} where id=?',(rid,))
    c.commit();c.close();flash(f'{len(ids)} case(s) processed','success');return redirect(url_for('cases',module=module))

@app.route('/collection/<int:rid>/ledger/<kind>',methods=['POST'])
@auth
def ledger(rid,kind):
    c=db()
    if kind=='payment':c.execute('insert into collection_payments(collection_case_id,payment_date,amount,payment_mode,receipt_number,remarks,created_at,created_by) values(?,?,?,?,?,?,?,?)',(rid,request.form.get('date'),request.form.get('amount') or 0,request.form.get('mode'),request.form.get('receipt'),request.form.get('remarks'),now(),session.get('username')))
    elif kind=='ptp':c.execute('insert into collection_ptp_ledger(collection_case_id,ptp_date,ptp_amount,outcome,remarks,created_at,created_by) values(?,?,?,?,?,?,?)',(rid,request.form.get('date'),request.form.get('amount') or 0,request.form.get('outcome'),request.form.get('remarks'),now(),session.get('username')))
    elif kind=='visit':c.execute('insert into collection_visits(collection_case_id,visit_date,visit_result,person_met,remarks,created_at,created_by) values(?,?,?,?,?,?,?)',(rid,request.form.get('date'),request.form.get('result'),request.form.get('person_met'),request.form.get('remarks'),now(),session.get('username')))
    c.commit();c.close();log('collection',rid,kind.title()+' Added');return redirect(url_for('detail',module='collection',rid=rid))

ALIASES={'collection':{'loan_number':['loan number','loan no','lan'],'customer_name':['customer name','borrower name'],'area':['area'],'district':['district'],'mobile':['mobile','mob no'],'loan_amount':['loan amount'],'emi':['emi'],'overdue_amount':['overdue amount'],'total_outstanding':['total outstanding'],'present_address':['present address'],'permanent_address':['permanent address'],'status':['status'],'remarks':['remarks']},'enforcement':{'case_number':['case number','case no'],'loan_number':['loan number','loan no','lan'],'borrower_name':['borrower name','customer name'],'village_city':['village','city'],'tehsil':['tehsil'],'district':['district'],'mobile':['mobile','mob no'],'bank_authorised_officer':['bank authorised officer'],'dm_order_status':['dm order status','dm orders'],'duty_magistrate':['duty magistrate'],'field_officer':['field officer'],'confirmation_date':['confirmation date','date of confirmation'],'affiliation_date':['affiliation date'],'cost_approval_status':['cost approval'],'administrative_expense':['administrative expense'],'possession_date':['possession date','date of possession'],'status':['status','possession status'],'remarks':['remarks']}}
def norm(x):return re.sub(r'[^a-z0-9]+',' ',str(x or '').strip().lower()).strip()
@app.route('/cases/<module>/import',methods=['POST'])
@auth
@mgr
def import_cases(module):
    f=request.files.get('file');bank=request.form.get('bank_id');
    if not f or not bank:flash('Choose Excel and Bank','warning');return redirect(url_for('cases',module=module))
    try:
        wb=load_workbook(f,data_only=True,read_only=True);ws=wb.active;rows=list(ws.iter_rows(values_only=True));aliases=ALIASES[module];best=0;score=-1
        for i,row in enumerate(rows[:30]):
            n=[norm(v) for v in row];s=sum(any(norm(a) in n for a in als) for als in aliases.values())
            if s>score:best,score=i,s
        heads=[norm(v) for v in rows[best]];mp={}
        for field,als in aliases.items():
            for a in als:
                if norm(a) in heads:mp[field]=heads.index(norm(a));break
        req=('loan_number','customer_name') if module=='collection' else ('loan_number','borrower_name')
        if any(k not in mp for k in req):raise ValueError('Required Loan Number/Borrower columns missing')
        t='collection_cases' if module=='collection' else 'enforcement_cases';c=db();added=updated=skipped=0
        for row in rows[best+1:]:
            d={k:(str(row[i]).strip() if i<len(row) and row[i] is not None else '') for k,i in mp.items()}
            if not any(d.values()):continue
            if any(not d.get(k) for k in req):skipped+=1;continue
            if module=='enforcement':d['case_number']=d.get('case_number') or d['loan_number']
            d['bank_id']=bank;d['status']=d.get('status') or ('Pending' if module=='collection' else 'Pending Status');d['created_at']=d['updated_at']=now();unique='loan_number' if module=='collection' else 'case_number';ex=c.execute(f'select id from {t} where {unique}=?',(d[unique],)).fetchone();keys=list(d)
            try:
                if ex:c.execute(f"update {t} set "+','.join(f'{k}=?' for k in keys if k!='created_at')+' where id=?',[d[k] for k in keys if k!='created_at']+[ex['id']]);updated+=1
                else:c.execute(f"insert into {t}({','.join(keys)}) values({','.join('?' for _ in keys)})",[d[k] for k in keys]);added+=1
            except:skipped+=1
        c.commit();c.close();flash(f'Import complete: {added} added, {updated} updated, {skipped} skipped','success')
    except Exception as e:flash('Import failed: '+str(e),'danger')
    return redirect(url_for('cases',module=module))
@app.route('/cases/<module>/export')
@auth
def export_cases(module):
    t='collection_cases' if module=='collection' else 'enforcement_cases';c=db();rows=c.execute(f'select * from {t} order by id desc').fetchall();c.close();out=io.StringIO();w=csv.writer(out);keys=list(rows[0].keys()) if rows else [];w.writerow(keys);[w.writerow([r[k] for k in keys]) for r in rows];return send_file(io.BytesIO(out.getvalue().encode('utf-8-sig')),mimetype='text/csv',as_attachment=True,download_name=f'Kredansh_{module}.csv')

MASTERS={'banks':['bank_name','contact_person','mobile','email','address','empanelment_no','agreement_date','status','remarks'],'employees':['employee_code','employee_name','designation','mobile','alternate_mobile','email','branch','area','district','department','status','date_of_joining','reporting_manager','remarks'],'directory':['state','district','tehsil','sub_tehsil','tehsildar_name','tehsildar_contact','reader_name','reader_contact','police_station_name','sho_name','sho_contact','police_station_contact','office_address','email','status','remarks']}
@app.route('/master/<name>')
@auth
def master(name):
    if name not in MASTERS:abort(404)
    table='tehsil_police_contacts' if name=='directory' else name;c=db();rows=c.execute(f'select * from {table} order by id desc').fetchall();c.close();return render_template('master.html',name=name,table=table,fields=MASTERS[name],rows=rows)
@app.route('/master/<name>/save',methods=['POST'])
@auth
@mgr
def master_save(name):
    table='tehsil_police_contacts' if name=='directory' else name;fields=MASTERS[name];rid=request.form.get('id');vals=[request.form.get(f,'') for f in fields];c=db()
    if rid:c.execute(f"update {table} set "+','.join(f'{f}=?' for f in fields)+' where id=?',vals+[rid])
    else:c.execute(f"insert into {table}({','.join(fields)}) values({','.join('?' for _ in fields)})",vals)
    c.commit();c.close();return redirect(url_for('master',name=name))

@app.route('/case/<module>/<int:rid>/documents',methods=['GET','POST'])
@auth
def documents(module,rid):
    folder=UPLOAD_DIR/'documents'/module/str(rid);folder.mkdir(parents=True,exist_ok=True)
    if request.method=='POST':
        f=request.files.get('file');
        if f and f.filename:f.save(folder/secure_filename(f.filename));log(module,rid,'Document Uploaded',f.filename)
        return redirect(url_for('documents',module=module,rid=rid))
    return render_template('documents.html',module=module,rid=rid,files=[p.name for p in folder.iterdir() if p.is_file()])
@app.route('/case/<module>/<int:rid>/documents/<name>')
@auth
def docfile(module,rid,name):return send_file(UPLOAD_DIR/'documents'/module/str(rid)/secure_filename(name))
@app.route('/case/<module>/<int:rid>/whatsapp')
@auth
def whatsapp(module,rid):
    t='collection_cases' if module=='collection' else 'enforcement_cases';n='customer_name' if module=='collection' else 'borrower_name';c=db();r=c.execute(f'select mobile,{n} name,status from {t} where id=?',(rid,)).fetchone();c.close();num=re.sub(r'\D','',r['mobile'] or '');num='91'+num if len(num)==10 else num;return redirect(f"https://wa.me/{num}?text={quote('Dear '+r['name']+', Kredansh Capital Services follow-up. Current status: '+str(r['status']))}")
@app.route('/case/<module>/<int:rid>/maps')
@auth
def maps(module,rid):
    t='collection_cases' if module=='collection' else 'enforcement_cases';f='present_address' if module=='collection' else 'property_address';c=db();r=c.execute(f'select {f} a from {t} where id=?',(rid,)).fetchone();c.close();return redirect('https://www.google.com/maps/search/?api=1&query='+quote(r['a'] or ''))

@app.route('/intelligence')
@auth
def intelligence():
    c=db();rows=[]
    for module,t,n,ref in [('collection','collection_cases','customer_name','loan_number'),('enforcement','enforcement_cases','borrower_name','case_number')]:
        for r in c.execute(f'select x.id,x.{n} borrower,x.{ref} ref,x.status,x.updated_at,b.bank_name,e.employee_name from {t} x left join banks b on b.id=x.bank_id left join employees e on e.id=x.employee_id').fetchall():
            stale=0
            try:stale=(datetime.date.today()-datetime.date.fromisoformat((r['updated_at'] or today())[:10])).days
            except:pass
            d=dict(r);d['module']=module;d['stale']=stale;d['score']=max(0,100-min(stale*3,50)-(20 if not r['employee_name'] else 0));rows.append(d)
    c.close();rows.sort(key=lambda x:x['score']);return render_template('intelligence.html',rows=rows)
@app.route('/daily-report')
@auth
def daily_report():
    d=request.args.get('date') or today();c=db();col=c.execute('select count(*) c,coalesce(sum(amount_collected),0) amount from collection_cases where collection_date=?',(d,)).fetchone();ptp=c.execute('select count(*) c from collection_cases where ptp_date=?',(d,)).fetchone();vis=c.execute('select count(*) c from collection_cases where visit_date=?',(d,)).fetchone();pos=c.execute('select count(*) c from enforcement_cases where possession_date=?',(d,)).fetchone();c.close();return render_template('daily.html',date=d,col=col,ptp=ptp,vis=vis,pos=pos)

@app.route('/users')
@auth
@mgr
def users():
    c=db();rows=c.execute('select * from users order by id').fetchall();c.close();return render_template('users.html',rows=rows,roles=ROLES)
@app.route('/users/save',methods=['POST'])
@auth
@mgr
def save_user():
    rid=request.form.get('id');u=request.form.get('username');full=request.form.get('full_name');role=request.form.get('role');status=request.form.get('status');pw=request.form.get('password');c=db()
    if rid:
        if pw:c.execute('update users set username=?,full_name=?,role=?,status=?,password=? where id=?',(u,full,role,status,generate_password_hash(pw),rid))
        else:c.execute('update users set username=?,full_name=?,role=?,status=? where id=?',(u,full,role,status,rid))
    else:c.execute('insert into users(username,password,full_name,role,status) values(?,?,?,?,?)',(u,generate_password_hash(pw or 'ChangeMe123'),full,role,status))
    c.commit();c.close();return redirect('/users')

try:init_db()
except Exception as e:print('DB INIT WARNING',e,flush=True)
if __name__=='__main__':app.run(host='0.0.0.0',port=int(os.environ.get('PORT','5050')))
